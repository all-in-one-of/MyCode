# -*- encoding: utf-8 -*-
"""
@File    : readExcel.py
@Time    : 2019/4/3 10:41
@Author  : Intime
@Software: PyCharm
"""
from openpyxl import load_workbook,workbook
from IntimeTool.usual import *

xlsx = r"F:\Share\原始模型\高居明作\高居明作 信息完成.xlsx"
xlsx1 = r"F:\Share\原始模型\高居明作\高居明作 信息完成1.xlsx"

wb = load_workbook(xlsx)

ws = wb['Sheet1']
ws1 = wb.create_sheet('已拍照')
ws2 = wb.create_sheet('未拍照')
ws3 = wb.create_sheet('不制作')

a = []
b = []
c = []
for i in ws.values:
    if i[-1] is 1:
        a.append(i)
    elif i[-1] is 2:
        b.append(i)
    else:
        c.append(i)
for i in a:
    ws1['A'+str(a.index(i)+1)]=i[1]
    ws1['B'+str(a.index(i)+1)]=i[3]

for i in b:
    ws2['A'+str(b.index(i)+1)]=i[1]
    ws2['B'+str(b.index(i)+1)]=i[3]

for i in c:
    ws3['A'+str(c.index(i)+1)]=i[1]
    ws3['B'+str(c.index(i)+1)]=i[3]

wb.save(xlsx1)